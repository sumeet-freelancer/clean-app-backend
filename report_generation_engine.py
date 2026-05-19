from datetime import datetime
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

SUMMARY_SHEET = "定期清掃報告書"
WEEK_SHEETS = ["1週目", "2週目", "3週目", "4週目"]
SUMMARY_DATE_COLUMNS = [5, 6, 7, 8]

WEEK_IMAGE_ANCHORS = {
    "1週目": ["B4", "B14", "N14", "A23", "N23", "B32", "N32"],
    "2週目": ["B4", "B14", "N14", "B22", "M23", "B32", "N32"],
    "3週目": ["B4", "B14", "N14", "A23", "N23", "N32", "B32"],
    "4週目": ["B4", "B14", "N14", "B23", "N23", "N32", "B32"],
}

BOTTOM_LABEL_CELLS = {
    "1週目": ("A31", "M31"),
    "2週目": ("A31", "M31"),
    "3週目": ("A31", "M31"),
    "4週目": ("A31", "M31"),
}


def make_reiwa_year_month(month_str: str) -> str:
    year, month = month_str.split("-")
    reiwa_year = int(year) - 2018
    return f"令和{reiwa_year}年{int(month)}月"


def safe_filename(value: str, fallback: str):
    normalized = re.sub(r"[^\w\-]+", "_", str(value or "").strip(), flags=re.ASCII).strip("_")
    return normalized or fallback


def make_report_date_text(work_date):
    if isinstance(work_date, datetime):
        dt = work_date
    else:
        dt = datetime.combine(work_date, datetime.min.time())
    return f"作業実施日：{dt.year}年　　{dt.month}月　　{dt.day}日"


def add_sheet_images(worksheet, image_paths):
    anchors = WEEK_IMAGE_ANCHORS.get(worksheet.title, [])
    for image_path, anchor in zip(image_paths[: len(anchors)], anchors):
        image = XLImage(str(image_path))
        worksheet.add_image(image, anchor)


def fill_summary_sheet(summary_ws, property_name, property_code, week_entries):
    summary_ws["B8"] = property_name
    if property_code:
        summary_ws["G8"] = f"物件コード：{property_code}"

    for index, entry in enumerate(week_entries[:4]):
        column = SUMMARY_DATE_COLUMNS[index]
        cell = summary_ws.cell(row=12, column=column)
        cell.value = entry["work_date"]
        cell.number_format = 'm"月"d"日"'


def fill_week_sheet(worksheet, property_name, week_entry):
    worksheet["N1"] = make_report_date_text(week_entry["work_date"])
    worksheet["D2"] = f"物件名称：{property_name}"

    left_label_cell, right_label_cell = BOTTOM_LABEL_CELLS.get(worksheet.title, ("A31", "M31"))
    if week_entry.get("bottom_left_label"):
        worksheet[left_label_cell] = week_entry["bottom_left_label"]
    if week_entry.get("bottom_right_label"):
        worksheet[right_label_cell] = week_entry["bottom_right_label"]

    add_sheet_images(worksheet, week_entry["photos"])


def generate_report(*, property_name: str, property_code: str, target_month: str, template_path: Path, output_dir: Path, week_entries):
    if not template_path.exists():
        raise FileNotFoundError(f"template not found: {template_path}")
    if not week_entries:
        raise RuntimeError("no week entries found for report generation")

    output_dir.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(template_path)

    summary_ws = workbook[SUMMARY_SHEET]
    fill_summary_sheet(summary_ws, property_name, property_code, week_entries)

    for sheet_name, entry in zip(WEEK_SHEETS, week_entries):
        if sheet_name not in workbook.sheetnames:
            continue
        fill_week_sheet(workbook[sheet_name], property_name, entry)

    property_token = safe_filename(property_code or property_name, "report")
    output_name = f"{target_month}_{property_token}_regular_cleaning_report.xlsx"
    output_path = output_dir / output_name
    workbook.save(output_path)
    return output_path
